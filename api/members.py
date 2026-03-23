import datetime
from fastapi import APIRouter, Depends, HTTPException, Response, status
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import List
from urllib.parse import quote

# 自作モジュール (同じ階層にある想定)
from . import models, database

router = APIRouter()


# --- Schemas ---
class MemberCreate(BaseModel):
    name: str
    role: str
    status: str
    skills: str | None = None
    price: int
    job_history: str | None = None
    qualifications: str | None = None
    evaluation: int = 3
    custom_fields: dict | None = {}


class MemberSchema(MemberCreate):
    id: int
    created_at: object
    updated_at: object

    class Config:
        from_attributes = True


class FieldSettingCreate(BaseModel):
    field_key: str
    label: str
    field_type: str = "text"


class FieldSettingSchema(FieldSettingCreate):
    id: int

    class Config:
        from_attributes = True


# --- Field Settings Endpoints ---

@router.get("/settings/fields", response_model = List[FieldSettingSchema])
def get_field_settings(db: Session = Depends(database.get_db)):
    return db.query(models.MemberFieldSetting).all()


@router.post("/settings/fields", response_model = FieldSettingSchema)
def create_field_setting(setting: FieldSettingCreate, db: Session = Depends(database.get_db)):
    db_setting = models.MemberFieldSetting(**setting.model_dump())
    db.add(db_setting)
    db.commit()
    db.refresh(db_setting)
    return db_setting


@router.delete("/settings/fields/{setting_id}")
def delete_field_setting(setting_id: int, db: Session = Depends(database.get_db)):
    db_setting = db.query(models.MemberFieldSetting).filter(models.MemberFieldSetting.id == setting_id).first()
    if db_setting:
        db.delete(db_setting)
        db.commit()
    return Response(status_code = status.HTTP_204_NO_CONTENT)


# --- Member Endpoints ---

@router.get("/members", response_model = List[MemberSchema])
def get_members(db: Session = Depends(database.get_db)):
    return db.query(models.Member).all()


@router.post("/members", response_model = MemberSchema)
def create_member(member: MemberCreate, db: Session = Depends(database.get_db)):
    db_member = models.Member(
        name = member.name,
        role = member.role,
        status = member.status,
        skills = member.skills,
        price = member.price,
        job_history = member.job_history,
        qualifications = member.qualifications,
        evaluation = member.evaluation,
        custom_fields = member.custom_fields
    )
    db.add(db_member)
    db.commit()
    db.refresh(db_member)
    return db_member


@router.put("/members/{member_id}", response_model = MemberSchema)
def update_member(member_id: int, member_in: MemberCreate, db: Session = Depends(database.get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code = 404, detail = "Member not found")

    db_member.name = member_in.name
    db_member.role = member_in.role
    db_member.status = member_in.status
    db_member.skills = member_in.skills
    db_member.price = member_in.price
    db_member.job_history = member_in.job_history
    db_member.qualifications = member_in.qualifications
    db_member.evaluation = member_in.evaluation
    db_member.custom_fields = member_in.custom_fields

    db.commit()
    db.refresh(db_member)
    return db_member


@router.delete("/members/{member_id}")
def delete_member(member_id: int, db: Session = Depends(database.get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code = 404, detail = "Member not found")

    db.delete(db_member)
    db.commit()
    return Response(status_code = status.HTTP_204_NO_CONTENT)


# --- Excel Download ---

@router.get("/members/{member_id}/download")
def download_skill_sheet(member_id: int, db: Session = Depends(database.get_db)):
    member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code = 404, detail = "Member not found")
    field_settings = db.query(models.MemberFieldSetting).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "職務経歴書"

    # Style definitions
    font_title = Font(name = 'ＭＳ ゴシック', size = 16, bold = True)
    font_header = Font(name = 'ＭＳ ゴシック', size = 11, bold = True)
    font_body = Font(name = 'ＭＳ 明朝', size = 11)
    border_thin = Side(border_style = "thin", color = "000000")
    border_all = Border(left = border_thin, right = border_thin, top = border_thin, bottom = border_thin)
    fill_header = PatternFill(start_color = "D9D9D9", end_color = "D9D9D9", fill_type = "solid")
    center_align = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    left_align = Alignment(horizontal = 'left', vertical = 'top', wrap_text = True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    ws.merge_cells('A1:E1')
    cell_title = ws['A1']
    cell_title.value = "職 務 経 歴 書"
    cell_title.font = font_title
    cell_title.alignment = center_align

    today_str = datetime.date.today().strftime("%Y年%m月%d日 現在")
    ws.merge_cells('D2:E2')
    ws['D2'] = today_str
    ws['D2'].alignment = Alignment(horizontal = 'right')

    current_row = 4

    def make_row(row, label1, val1, label2 = None, val2 = None):
        c1 = ws.cell(row = row, column = 1, value = label1)
        c1.font = font_header
        c1.fill = fill_header
        c1.border = border_all
        c1.alignment = center_align

        c2 = ws.cell(row = row, column = 2, value = val1)
        c2.font = font_body
        c2.border = border_all
        c2.alignment = left_align

        if label2:
            c3 = ws.cell(row = row, column = 3, value = label2)
            c3.font = font_header
            c3.fill = fill_header
            c3.border = border_all
            c3.alignment = center_align

            c4 = ws.cell(row = row, column = 4, value = val2)
            c4.font = font_body
            c4.border = border_all
            c4.alignment = left_align

            ws.merge_cells(start_row = row, start_column = 4, end_row = row, end_column = 5)
            for col in range(4, 6):
                ws.cell(row = row, column = col).border = border_all
        else:
            ws.merge_cells(start_row = row, start_column = 2, end_row = row, end_column = 5)
            for col in range(2, 6):
                ws.cell(row = row, column = col).border = border_all

    eval_map = {1: "見習い", 2: "初級", 3: "普通", 4: "優秀", 5: "プロフェッショナル"}
    eval_str = eval_map.get(member.evaluation, "普通")

    make_row(current_row, "氏名", f"{member.name}  殿", "役割", member.role)
    current_row += 1
    make_row(current_row, "性別", "（非公開）", "年齢", "（非公開）")
    current_row += 1
    make_row(current_row, "最寄駅", "（非公開）", "稼働状況", member.status)
    current_row += 1
    make_row(current_row, "社内評価", eval_str, "保有資格", member.qualifications or "なし")
    current_row += 1

    custom_data = member.custom_fields or {}
    # カスタム項目の出力ループ修正
    for i in range(0, len(field_settings), 2):
        fs1 = field_settings[i]
        val1 = custom_data.get(fs1.field_key, "")
        fs2 = None
        val2 = None
        if i + 1 < len(field_settings):
            fs2 = field_settings[i + 1]
            val2 = custom_data.get(fs2.field_key, "")

        make_row(current_row, fs1.label, val1, fs2.label if fs2 else None, val2)
        current_row += 1

    make_row(current_row, "技術スキル", member.skills or "なし")
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    ws.merge_cells(start_row = current_row, start_column = 1, end_row = current_row, end_column = 5)
    header_history = ws.cell(row = current_row, column = 1, value = "【 職務経歴詳細 】")
    header_history.font = font_header
    header_history.fill = fill_header
    header_history.border = border_all
    current_row += 1

    history_start_row = current_row
    history_end_row = current_row + 20

    ws.merge_cells(start_row = history_start_row, start_column = 1, end_row = history_end_row, end_column = 5)
    cell_history = ws.cell(row = history_start_row, column = 1, value = member.job_history or "（職務経歴の登録がありません）")
    cell_history.font = font_body
    cell_history.alignment = Alignment(horizontal = 'left', vertical = 'top', wrap_text = True)

    for r in range(history_start_row, history_end_row + 1):
        for c in range(1, 6):
            ws.cell(row = r, column = c).border = border_all

    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"職務経歴書_{member.name}.xlsx"
    encoded_filename = quote(filename)

    headers = {
        'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
    }

    return StreamingResponse(output, headers = headers, media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
